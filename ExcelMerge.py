import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook

# ==================================================
# PAGE CONFIG
# ==================================================

st.set_page_config(
    page_title="Atlas Copco Excel Merger",
    layout="wide"
)

st.title("Atlas Copco Excel Merger")

st.write("""
Upload multiple Registered Product reports.

The app will:

✅ Automatically find the header row

✅ Remove report metadata

✅ Merge files

✅ Remove duplicate IDs

✅ Download a clean merged workbook
""")

# ==================================================
# FUNCTIONS
# ==================================================

def load_excel(uploaded_file):
    """
    Uses openpyxl directly instead of pandas.read_excel.
    This is more tolerant of problematic Atlas Copco exports.
    """

    uploaded_file.seek(0)

    file_bytes = BytesIO(uploaded_file.read())

    workbook = load_workbook(
        filename=file_bytes,
        data_only=True,
        read_only=True
    )

    sheets = {}

    for ws in workbook.worksheets:

        rows = []

        for row in ws.iter_rows(values_only=True):
            rows.append(list(row))

        sheets[ws.title] = pd.DataFrame(rows)

    return sheets


def clean_sheet(df):
    """
    Locate header row containing ID
    """

    header_row = None

    for i in range(len(df)):

        values = [
            str(v).strip()
            for v in df.iloc[i].fillna("")
        ]

        if "ID" in values:
            header_row = i
            break

    if header_row is None:
        return None

    headers = df.iloc[header_row]

    cleaned_df = df.iloc[header_row + 1:].copy()

    cleaned_df.columns = headers

    cleaned_df = cleaned_df.reset_index(drop=True)

    cleaned_df = cleaned_df.dropna(how="all")

    return cleaned_df


# ==================================================
# USER SETTINGS
# ==================================================

uploaded_files = st.file_uploader(
    "Upload Excel Files",
    type=["xlsx", "xls"],
    accept_multiple_files=True
)

remove_duplicates = st.checkbox(
    "Remove Duplicate IDs",
    value=True
)

show_preview = st.checkbox(
    "Show Preview",
    value=True
)

# ==================================================
# PROCESS
# ==================================================

if st.button("Process Files"):

    if not uploaded_files:

        st.warning("Please upload one or more Excel files.")

    else:

        merged_frames = []
        failed_files = []

        progress_bar = st.progress(0)

        for index, uploaded_file in enumerate(uploaded_files):

            try:

                workbook = load_excel(uploaded_file)

                for sheet_name, df in workbook.items():

                    try:

                        cleaned_df = clean_sheet(df)

                        if cleaned_df is None:
                            continue

                        cleaned_df["Source File"] = (
                            uploaded_file.name
                        )

                        merged_frames.append(cleaned_df)

                    except Exception as sheet_error:

                        failed_files.append(
                            f"{uploaded_file.name} | "
                            f"{sheet_name} | "
                            f"{sheet_error}"
                        )

            except Exception as file_error:

                failed_files.append(
                    f"{uploaded_file.name} | "
                    f"{file_error}"
                )

            progress_bar.progress(
                (index + 1) / len(uploaded_files)
            )

        # ==========================================
        # NO DATA
        # ==========================================

        if not merged_frames:

            st.error(
                "No valid sheets could be processed."
            )

        else:

            # ==========================================
            # MERGE
            # ==========================================

            final_df = pd.concat(
                merged_frames,
                ignore_index=True
            )

            # ==========================================
            # CLEAN COLUMN NAMES
            # ==========================================

            clean_columns = []

            for i, col in enumerate(final_df.columns):

                if pd.isna(col):

                    clean_columns.append(
                        f"Column_{i}"
                    )

                else:

                    clean_columns.append(
                        str(col)
                    )

            final_df.columns = clean_columns

            final_df = final_df.fillna("")

            # ==========================================
            # REMOVE DUPLICATES
            # ==========================================

            if remove_duplicates:

                if "ID" in final_df.columns:

                    before_count = len(final_df)

                    final_df = final_df.drop_duplicates(
                        subset=["ID"]
                    )

                    removed = (
                        before_count
                        - len(final_df)
                    )

                    st.success(
                        f"Removed {removed:,} duplicates."
                    )

            # ==========================================
            # SUMMARY
            # ==========================================

            st.success(
                f"Merged {len(uploaded_files)} files."
            )

            st.success(
                f"Final row count: {len(final_df):,}"
            )

            if failed_files:

                st.warning(
                    f"{len(failed_files)} file(s) or sheet(s) failed."
                )

                with st.expander(
                    "View Failed Files"
                ):

                    for item in failed_files:
                        st.write(item)

            # ==========================================
            # PREVIEW
            # ==========================================

            if show_preview:

                st.subheader("Preview")

                preview_df = (
                    final_df
                    .head(100)
                    .fillna("")
                    .astype(str)
                )

                st.write(
                    f"Showing first {len(preview_df)} rows "
                    f"of {len(final_df):,}"
                )

                st.dataframe(
                    preview_df,
                    use_container_width=True
                )

            # ==========================================
            # CREATE OUTPUT FILE
            # ==========================================

            output = BytesIO()

            with pd.ExcelWriter(
                output,
                engine="openpyxl"
            ) as writer:

                final_df.to_excel(
                    writer,
                    sheet_name="Merged_Data",
                    index=False
                )

            output.seek(0)

            st.download_button(
                label="📥 Download Merged Excel File",
                data=output.getvalue(),
                file_name="Merged_Registered_Products.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

# ==================================================
# FOOTER
# ==================================================

st.divider()

st.caption(
    "Atlas Copco Excel Merger"
)

st.markdown('### GitHub Requirements')
st.code('''
streamlit
pandas
openpyxl
xlrd
''', language='text')
