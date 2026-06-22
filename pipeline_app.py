import streamlit as st
import pandas as pd
import re
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from rapidfuzz import fuzz

# ------------------ UI CONFIG ------------------
st.set_page_config(page_title="Dedup Pipeline Tool", layout="wide")

st.title("🔍 Excel Deduplication Tool")
st.markdown("### Upload → Process → Download (CA-style output)")

with st.expander("ℹ️ How this works"):
    st.write("""
    - Removes metadata rows + first column
    - Normalizes names & addresses
    - Uses fuzzy matching to group duplicates
    - Outputs:
        - Highlighted main sheet
        - Duplicate groups
        - Merge instructions
        - Dashboard
    """)

uploaded_file = st.file_uploader("📂 Upload Excel file", type=["xlsx"])

# ------------------ NORMALIZATION ------------------
def normalize(s):
    if pd.isna(s):
        return ""
    s = str(s).lower()
    s = re.sub(r'[^a-z0-9 ]', ' ', s)
    s = re.sub(r'\s+', ' ', s).strip()
    return s

# ------------------ FUZZY MATCH ------------------
def fuzzy_match(row1, row2):
    name_score = fuzz.token_sort_ratio(row1["norm_name"], row2["norm_name"])
    addr_score = fuzz.token_sort_ratio(row1["norm_address"], row2["norm_address"])

    return name_score >= 90 and addr_score >= 90

# ------------------ PIPELINE ------------------
# Normalize column names
df_raw.columns = [str(col).strip() for col in df_raw.columns]

# Try to auto-detect key columns
def find_column(possible_names):
    for col in df_raw.columns:
        if any(name.lower() in col.lower() for name in possible_names):
            return col
    return None

name_col = find_column(["name"])
address_col = find_column(["address"])
state_col = find_column(["state"])

def run_pipeline(df):
    # Remove first column
    df = df.iloc[:, 1:].reset_index(drop=True)

    # Normalize
    if not name_col or not address_col:
        st.error("❌ Could not detect Name or Address columns. Please check file format.")
        st.stop()
    
    df["norm_name"] = df[name_col].fillna("").astype(str).str.lower()
    df["norm_address"] = df[address_col].fillna("").astype(str).str.lower()

    # ---------- BLOCKING ----------
    df["block"] = (
        df["norm_name"].str[:3] + "_" +
        df["norm_address"].str[:5]
    )

    df["group_id"] = None
    df["match_score"] = None

    group_counter = 1

    # ---------- MATCHING ----------
    for block, subset in df.groupby("block"):
        idxs = subset.index.tolist()

        for i_idx in range(len(idxs)):
            i = idxs[i_idx]

            if pd.notna(df.loc[i, "group_id"]):
                continue

            group = [i]

            for j_idx in range(i_idx + 1, len(idxs)):
                j = idxs[j_idx]

                if pd.notna(df.loc[j, "group_id"]):
                    continue

                # Fuzzy scores
                name_score = fuzz.token_sort_ratio(
                    df.loc[i, "norm_name"],
                    df.loc[j, "norm_name"]
                )

                addr_score = fuzz.token_sort_ratio(
                    df.loc[i, "norm_address"],
                    df.loc[j, "norm_address"]
                )

                final_score = 0.6 * name_score + 0.4 * addr_score

                # Decision thresholds
                if final_score >= 92:
                    group.append(j)
                    df.loc[j, "match_score"] = final_score

            if len(group) > 1:
                for idx in group:
                    df.loc[idx, "group_id"] = group_counter
                    if pd.isna(df.loc[idx, "match_score"]):
                        df.loc[idx, "match_score"] = 100

                group_counter += 1

    # ---------- DUP COUNT ----------
    df["dup_count"] = df["group_id"].map(df["group_id"].value_counts())

    # ---------- MASTER SELECTION ----------
    def pick_master(group):
        # Prefer non-Prospect
        if "Type" in group.columns:
            group = group.sort_values(by="Type", ascending=False)

        # Prefer longest name (more complete)
        return group.iloc[group["Name"].str.len().idxmax()]

    actions = []

    for gid, group in df.groupby("group_id", dropna=False):
        if pd.isna(gid):
            for idx in group.index:
                actions.append((idx, "UNIQUE"))
        else:
            master_idx = group["Name"].str.len().idxmax()

            for idx in group.index:
                if idx == master_idx:
                    actions.append((idx, "KEEP (Master)"))
                else:
                    actions.append((idx, f"MERGE into {master_idx}"))

    df["Action"] = pd.Series(dict(actions))

    # ---------- SORT ----------
    dups = df[df["group_id"].notna()].sort_values(
        ["dup_count", "group_id"], ascending=[False, True]
    )
    singles = df[df["group_id"].isna()]

    return pd.concat([dups, singles])

# ------------------ EXCEL OUTPUT ------------------
def build_file(df):
    output = BytesIO()
    wb = Workbook()

    ws_main = wb.active
    ws_main.title = "Main"

    ws_main.append(list(df.columns))
    for row in df.itertuples(index=False):
        ws_main.append(list(row))

    # Colors
    colors = ["FFFF99","99FF99","99CCFF","FF9999","CC99FF"]
    color_map = {}
    idx = 0

    for gid in df["group_id"].dropna().unique():
        color_map[gid] = colors[idx % len(colors)]
        idx += 1

    for i, row in enumerate(df.itertuples(index=False), start=2):
        gid = row.group_id
        if pd.notna(gid):
            fill = PatternFill(start_color=color_map[gid],
                               end_color=color_map[gid],
                               fill_type="solid")
            for c in range(1, len(df.columns)+1):
                ws_main.cell(row=i, column=c).fill = fill

    # Duplicate Groups
    ws_groups = wb.create_sheet("Duplicate Groups")

    for gid, group in df[df["group_id"].notna()].groupby("group_id"):
        ws_groups.append([f"==== GROUP {int(gid)} ===="])
        ws_groups.append(list(df.columns))
        for r in group.itertuples(index=False):
            ws_groups.append(list(r))
        ws_groups.append([])

    # Merge Sheet
    ws_merge = wb.create_sheet("Merge Instructions")
    ws_merge.append(["group_id","Action"])

    for r in df[df["Action"].str.contains("MERGE", na=False)].itertuples(index=False):
        ws_merge.append([r.group_id, r.Action])

    # Summary
    ws_summary = wb.create_sheet("Summary Dashboard")
    ws_summary.append(["Group ID","Count"])

    summary = df[df["group_id"].notna()].groupby("group_id").size().sort_values(ascending=False)

    for gid, count in summary.items():
        ws_summary.append([gid, count])

    wb.save(output)
    return output

# ------------------ RUN ------------------
uploaded_file = st.file_uploader("📂 Upload Excel file", type=["xlsx"])

if uploaded_file:

    st.success("✅ File uploaded")

    # ✅ Step 1: Read file FIRST
    df_raw = safe_read_excel(uploaded_file)

    # ✅ Step 2: THEN clean columns
    df_raw.columns = [str(col).strip() for col in df_raw.columns]

    # ✅ Step 3: Detect columns
    def find_column(possible_names):
        for col in df_raw.columns:
            if any(name.lower() in col.lower() for name in possible_names):
                return col
        return None

    name_col = find_column(["name"])
    address_col = find_column(["address"])
    state_col = find_column(["state"])

    if not name_col or not address_col:
        st.error("❌ Could not detect Name or Address columns.")
        st.stop()

    st.success(f"✅ Using columns: {name_col}, {address_col}")

    # ✅ Optional preview
    st.dataframe(df_raw.head())

    if st.button("🚀 Run Pipeline"):
        with st.spinner("Processing..."):
            result = run_pipeline(df_raw)

            excel_file = build_file(result)

        st.success("✅ Pipeline Complete!")

        # ✅ Dynamic file name
        if state_col:
            state_name = result[state_col].dropna().astype(str).str.title().mode()[0]
        else:
            state_name = "output"

        file_name = f"{state_name.replace(' ','_')}_final_processed.xlsx"

        st.download_button(
            "⬇ Download Output",
            excel_file.getvalue(),
            file_name=file_name
        )

def safe_read_excel(file):
   try:
       df = pd.read_excel(file, skiprows=4, engine="openpyxl")
   except Exception:
       file.seek(0)
       df = pd.read_excel(file, header=None, engine="openpyxl")
   
   # force columns to strings
   df.columns = df.columns.astype(str)
   return df

df_raw = safe_read_excel(uploaded_file)

st.write("Preview:")
st.dataframe(df_raw.head())

if st.button("🚀 Run Pipeline"):
    with st.spinner("Processing..."):
        result = run_pipeline(df_raw)

        excel_file = build_file(result)

    st.success("✅ Pipeline Complete!")

# Detect state name
if "State" in result.columns:
    state_name = result["State"].dropna().mode()[0]  # most frequent state
else:
    state_name = "output"

# Clean state name (safe filename)
state_name_clean = str(state_name).strip().replace(" ", "_")

file_name = f"{state_name_clean}_final_processed.xlsx"

st.download_button(
    "⬇ Download Output",
    excel_file.getvalue(),
    file_name=file_name
)
