import streamlit as st
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill
import io
import time

st.set_page_config(page_title="Duplicate Row Auditor", layout="wide")
st.title("📊 Duplicate Row Auditor")

uploaded_files = st.file_uploader(
    "Upload Excel files",
    type=["xlsx"],
    accept_multiple_files=True
)

# -------------------------------
# ✅ HELPERS
# -------------------------------

def normalize(col):
    if not col:
        return ""
    return (
        str(col)
        .strip()
        .lower()
        .replace(" ", "")
        .replace("_", "")
        .replace("\xa0", "")
    )

def smart_column_map(headers):
    normalized = [normalize(h) for h in headers]

    def find_best(keywords):
        for i, col in enumerate(normalized):
            for k in keywords:
                if k in col:
                    return i
        return None

    return {
        "group_id": find_best(["groupid", "group"]),
        "account_group": find_best(["accountgroup"]),
        "external_id": find_best(["externalid", "external", "extid", "id"])
    }

def is_sold_to(val):
    return val and "sold to" in str(val).lower()

def is_highlighted(row):
    return any(cell.fill and cell.fill.fill_type for cell in row)

def safe_copy_fill(cell):
    try:
        if not cell.fill or not cell.fill.fill_type:
            return None
        return PatternFill(
            start_color=cell.fill.start_color.rgb,
            end_color=cell.fill.end_color.rgb,
            fill_type=cell.fill.fill_type
        )
    except:
        return None

# 🎨 Rainbow palette
COLOR_PALETTE = [
    "FFCCCC", "CCE5FF", "D5F5E3", "FFF2CC",
    "E8DAEF", "FADBD8", "D6EAF8", "FCF3CF"
]

# -------------------------------
# ✅ PROCESS FILES
# -------------------------------

def process_files(file_dict_list):

    all_0, all_1, all_2 = [], [], []
    summary_stats = []
    error_files = []
    headers = None

    for file_dict in file_dict_list:

        filename = file_dict["name"]
        file_bytes = file_dict["bytes"]

        try:
            wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
            ws = wb.active
        except:
            error_files.append(filename)
            continue

        headers_raw = [cell.value for cell in ws[1]]
        col_map = smart_column_map(headers_raw)

        if col_map["group_id"] is None or col_map["account_group"] is None:
            error_files.append(filename)
            continue

        if headers is None:
            headers = ["Source File"] + headers_raw

        group_idx = col_map["group_id"]
        acc_idx = col_map["account_group"]
        ext_idx = col_map["external_id"]

        groups = {}

        for row in ws.iter_rows(min_row=2):

            if not is_highlighted(row):
                continue

            key = f"{filename}__{row[group_idx].value}"
            groups.setdefault(key, []).append(row)

        g0 = g1 = g2 = 0

        for group_rows in groups.values():

            sold_to_count = sum(
                1 for r in group_rows if is_sold_to(r[acc_idx].value)
            )

            if sold_to_count >= 2:
                all_2.append([(filename, r) for r in group_rows])
                g2 += 1
            elif sold_to_count == 1:
                all_1.append([(filename, r) for r in group_rows])
                g1 += 1
            else:
                all_0.append([(filename, r) for r in group_rows])
                g0 += 1

        summary_stats.append({
            "file": filename,
            "total": len(groups),
            "0": g0,
            "1": g1,
            "2": g2
        })

    return headers, all_0, all_1, all_2, summary_stats, error_files


# -------------------------------
# ✅ UI
# -------------------------------

if uploaded_files:

    if st.button("⚡ Generate Merged Excel"):

        progress_bar = st.progress(0)
        status_text = st.empty()

        file_dict_list = []
        total_files = len(uploaded_files)

        for i, file in enumerate(uploaded_files):
            status_text.text(f"Loading {file.name}...")
            file_dict_list.append({
                "name": file.name,
                "bytes": file.getvalue()
            })
            progress_bar.progress((i + 1) / total_files * 0.3)

        status_text.text("Processing duplicate groups...")

        headers, all_0, all_1, all_2, summary, errors = process_files(file_dict_list)

        progress_bar.progress(0.8)

        # -------------------------------
        # ✅ EXPORT
        # -------------------------------

        wb = Workbook()
        wb.remove(wb.active)

        def write_sheet(name, groups):

            ws = wb.create_sheet(name)

            if headers:
                for i, h in enumerate(headers, start=1):
                    ws.cell(1, i, h)

            row_cursor = 2

            for group in groups:

                # ✅ Build External ID rainbow map
                ext_map = {}
                color_index = 0

                for (filename, row) in group:
                    if len(row) == 0:
                        continue

                    if ext_idx := next(
                        (i for i, h in enumerate(headers[1:]) if "external" in normalize(h)), 
                        None
                    ):
                        ext_value = row[ext_idx].value
                        if ext_value and ext_value not in ext_map:
                            ext_map[ext_value] = COLOR_PALETTE[color_index % len(COLOR_PALETTE)]
                            color_index += 1

                for filename, row in group:

                    ws.cell(row_cursor, 1, filename)

                    ext_value = None
                    if ext_idx is not None:
                        ext_value = row[ext_idx].value

                    row_color = ext_map.get(ext_value)

                    for j, cell in enumerate(row, start=2):

                        new_cell = ws.cell(row_cursor, j, cell.value)

                        # preserve original highlight
                        base_fill = safe_copy_fill(cell)
                        if base_fill:
                            new_cell.fill = base_fill

                        # apply rainbow color
                        if row_color:
                            new_cell.fill = PatternFill(
                                start_color=row_color,
                                end_color=row_color,
                                fill_type="solid"
                            )

                    row_cursor += 1

                row_cursor += 1

        write_sheet("2+ SoldTo Accounts", all_2)
        write_sheet("1 SoldTo Accounts", all_1)
        write_sheet("0 SoldTo Accounts", all_0)

        # ✅ Summary sheet
        ws_summary = wb.create_sheet("Summary")
        ws_summary.append(["File", "Total Groups", "0", "1", "2+"])

        for s in summary:
            ws_summary.append([s["file"], s["total"], s["0"], s["1"], s["2"]])

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        progress_bar.progress(1.0)
        status_text.text("✅ Complete!")

        if errors:
            st.warning("⚠️ Some files skipped:")
            for e in errors:
                st.write(f"❌ {e}")

        st.success("✅ Excel generated successfully!")

        st.download_button(
            "⬇️ Download Merged Excel",
            data=output,
            file_name="Multiple Sold To Duplicates.xlsx"
        )
